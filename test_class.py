import io
from openpyxl import load_workbook

class File_Interact():
    def __init__(self,file_name):
        self.file_name=file_name

    def write_file(self,ndung):
        f = io.open(self.file_name,'w',encoding='utf-8')
        f.write(ndung)
        f.close()

    def write_file_from_list(self, list_lines):
        f = io.open(self.file_name, 'w', encoding='utf-8')
        f.write('\n'.join(list_lines))
        f.close()

    def write_file_line(self, ndung_line):
        f = io.open(self.file_name, 'a', encoding='utf-8')
        f.write('%s\n' % ndung_line)
        f.close()

    def read_file(self):
        f = io.open(self.file_name, 'r', encoding='utf-8')
        ndung = f.read()
        f.close()
        return ndung

    def read_file_list(self):
        f = io.open(self.file_name, 'r', encoding='utf-8')
        ndung = f.read()
        f.close()
        return ndung.split('\n')

class File_Excel():

    def __init__(self, file_name):
        self.file_name = file_name
    def read_cell(file_path, sheetname, cell_name):
        wb = load_workbook(filename=file_path)
        sheet_ranges = wb[sheetname]
        return sheet_ranges[cell_name].value
    def update_cell(file_path, sheetname, cell_name, new_value):
        wb = load_workbook(filename=file_path)
        wb[sheetname][cell_name].value = new_value
        wb.close()
        wb.save(file_path)
